"""XLSX Tabular leaf with lazy per-sheet entries.

:class:`XlsxIO` IS-A :class:`BytesIO` whose backing bytes are an
xlsx workbook (a ZIP archive). Mirrors the :class:`ZipIO` shape:

1. **Byte surface** — inherited from :class:`BytesIO`. Read / write
   the raw workbook bytes (e.g. when handing the workbook off to
   :mod:`openpyxl` directly).
2. **Children surface** — :meth:`iter_children` walks every
   worksheet as a :class:`XlsxSheetIO`. Sheets are **lazy**: rows
   are pulled out of the parent workbook on first read, not at
   directory-walk time.

Reads use :func:`openpyxl.load_workbook(read_only=True)`; writes
use :func:`openpyxl.Workbook(write_only=True)`.

Mode dispatch on a workbook write:

- **OVERWRITE / AUTO / TRUNCATE** — fresh workbook with one sheet
  whose data is the incoming batches; sheet name from
  ``options.sheet_name``.
- **APPEND** — pull the existing workbook's other sheets through
  openpyxl, drop any sheet whose name matches ``options.sheet_name``,
  then write a fresh workbook containing the survivors plus the
  new sheet.
- **IGNORE** — skip when non-empty.
- **ERROR_IF_EXISTS** — raise when non-empty.
- **UPSERT / MERGE** — degrade to APPEND.

Convenience helper :meth:`write_sheets` packs a ``{name: arrow_table}``
mapping into a fresh workbook.
"""

from __future__ import annotations

import dataclasses
import io as _stdlib_io
from datetime import date, datetime
from typing import TYPE_CHECKING, Any, ClassVar, Iterable, Iterator, Mapping

import pyarrow as pa

from yggdrasil.data.options import CastOptions
from yggdrasil.data.schema import Schema
from yggdrasil.data.enums import MimeTypes, Mode
from yggdrasil.io.bytes_io import BytesIO
from yggdrasil.io.memory import Memory

if TYPE_CHECKING:
    pass

__all__ = ["XlsxIO", "XlsxOptions", "XlsxSheetIO"]


#: Default sentinel strings that map to ``None`` during type inference.
#: Mirrors :attr:`yggdrasil.io.primitive.csv_io.CsvOptions.null_values`
#: with a couple of common spreadsheet markers added (``#N/A``,
#: ``-``, ``--``).
_DEFAULT_NULL_VALUES: tuple[str, ...] = (
    "", "N/A", "n/a", "NA", "na", "#N/A",
    "NULL", "null", "Null",
    "None", "none",
    "-", "--",
)

#: Date formats tried in order during string-cell inference. The
#: first match wins, so ISO comes first.
_DEFAULT_DATE_FORMATS: tuple[str, ...] = (
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%d-%m-%Y",
    "%d.%m.%Y",
)

#: Datetime formats. Tried before the date formats above on the same
#: input; a successful match returns a :class:`datetime.datetime`.
_DEFAULT_DATETIME_FORMATS: tuple[str, ...] = (
    "%Y-%m-%dT%H:%M:%S.%f",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d %H:%M:%S.%f",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%d/%m/%Y %H:%M:%S",
    "%m/%d/%Y %H:%M:%S",
)


@dataclasses.dataclass(frozen=True, slots=True)
class XlsxOptions(CastOptions):
    """:class:`CastOptions` extended with XLSX-specific knobs."""

    #: Active worksheet for single-sheet reads / writes. On read,
    #: missing names fall back to the first sheet; on write, this is
    #: the sheet that receives the incoming batches.
    sheet_name: str = "Sheet1"
    has_header: bool = True

    # ------------------------------------------------------------------
    # Read-side type inference (string cells only — openpyxl already
    # types numbers, dates, and booleans natively).
    # ------------------------------------------------------------------

    #: Master switch. When ``False``, raw cell values flow through
    #: untouched (openpyxl-native types for typed cells, plain
    #: strings for everything else).
    infer_types: bool = True

    #: String tokens that map to ``None`` during inference. Compared
    #: against the cell's stripped value (case-sensitive — list both
    #: cases when needed).
    null_values: "tuple[str, ...]" = _DEFAULT_NULL_VALUES

    #: Character treated as a thousands grouping separator inside
    #: numeric strings. ``"123,892.50"`` → ``123892.50``. Set to
    #: empty string to disable.
    thousands_separator: str = ","

    #: Decimal mark inside numeric strings. ``"1.234,56"`` parses
    #: when ``thousands_separator=","`` and ``decimal_separator=","``
    #: — but not both at once on the same character; the helper
    #: normalizes to ``.`` before :func:`float`.
    decimal_separator: str = "."

    #: ``strptime`` patterns for date inference. The first one that
    #: matches wins.
    date_formats: "tuple[str, ...]" = _DEFAULT_DATE_FORMATS

    #: ``strptime`` patterns for datetime inference. Tried before the
    #: date formats so an ``"YYYY-MM-DD HH:MM:SS"`` cell becomes a
    #: ``datetime``, not a ``date``.
    datetime_formats: "tuple[str, ...]" = _DEFAULT_DATETIME_FORMATS

    #: Column names whose cells should be left alone (no numeric /
    #: date coercion). Useful for IDs that look numeric but must
    #: stay string. Null-marker handling still applies.
    string_columns: "tuple[str, ...]" = ()


def _openpyxl():
    """Lazy openpyxl import shared by :class:`XlsxIO` and :class:`XlsxSheetIO`."""
    try:
        import openpyxl
    except ImportError as e:  # pragma: no cover
        raise ImportError(
            "XlsxIO requires openpyxl. Install with: pip install openpyxl"
        ) from e
    return openpyxl


def _rows_to_batches(rows: list[dict]) -> Iterator[pa.RecordBatch]:
    if not rows:
        return
    yield from pa.Table.from_pylist(rows).to_batches()


def _chain_one(first, rest):
    yield first
    yield from rest


def _coerce_cell(
    value: Any,
    *,
    null_values: frozenset,
    thousands_sep: str,
    decimal_sep: str,
    date_formats: tuple[str, ...],
    datetime_formats: tuple[str, ...],
    keep_string: bool,
) -> Any:
    """Apply the read-side inference rules to a single openpyxl cell.

    Non-string cells (openpyxl already typed them: int / float /
    bool / datetime / date / None) flow through unchanged. String
    cells walk this ladder:

    1. Strip whitespace.
    2. Empty or in *null_values* → ``None``.
    3. *keep_string* short-circuits the rest — caller asked for a
       string column (e.g. an ID).
    4. Numeric: strip *thousands_sep*, normalize *decimal_sep* to
       ``.``; try :func:`int`, then :func:`float`.
    5. Datetime: try each *datetime_formats* pattern.
    6. Date: try each *date_formats* pattern.
    7. Fall back to the original (un-stripped) string.
    """
    if value is None:
        return None
    if not isinstance(value, str):
        return value

    stripped = value.strip()
    if not stripped or stripped in null_values:
        return None
    if keep_string:
        return value

    # Numeric — handle a leading sign, optional thousands grouping,
    # optional non-period decimal mark.
    candidate = stripped
    if thousands_sep and thousands_sep != decimal_sep:
        candidate = candidate.replace(thousands_sep, "")
    if decimal_sep != ".":
        candidate = candidate.replace(decimal_sep, ".")
    if candidate and candidate[0] in "+-":
        sign, body = candidate[0], candidate[1:]
    else:
        sign, body = "", candidate
    if body and (body.replace(".", "", 1).isdigit() or _is_scientific(body)):
        try:
            if "." not in body and "e" not in body.lower():
                return int(sign + body)
        except ValueError:
            pass
        try:
            return float(sign + body)
        except ValueError:
            pass

    # Datetime first (more specific), then date.
    for fmt in datetime_formats:
        try:
            return datetime.strptime(stripped, fmt)
        except ValueError:
            continue
    for fmt in date_formats:
        try:
            return datetime.strptime(stripped, fmt).date()
        except ValueError:
            continue

    return value


def _is_scientific(body: str) -> bool:
    """``"1e5"`` / ``"2.5E-3"`` shape — mantissa + ``e`` + signed exponent."""
    lower = body.lower()
    if "e" not in lower:
        return False
    mantissa, _, exponent = lower.partition("e")
    if not mantissa or not exponent:
        return False
    if exponent[0] in "+-":
        exponent = exponent[1:]
    return (
        mantissa.replace(".", "", 1).isdigit()
        and exponent.isdigit()
    )


def _stream_sheet_batches(
    ws: Any,
    *,
    options: "XlsxOptions",
) -> Iterator[pa.RecordBatch]:
    """Yield arrow batches for *ws* in ``options.row_size`` chunks.

    Shared by :class:`XlsxIO` (single-sheet read) and
    :class:`XlsxSheetIO` (per-child read) so the column-discovery /
    batching / inference rules don't drift between them.
    """
    row_iter = ws.iter_rows(values_only=True)

    if options.has_header:
        headers = next(row_iter, None)
        if headers is None:
            return
        columns = [
            str(h) if h is not None else f"col_{i}"
            for i, h in enumerate(headers)
        ]
    else:
        first_row = next(row_iter, None)
        if first_row is None:
            return
        columns = [f"col_{i}" for i in range(len(first_row))]
        row_iter = _chain_one(first_row, row_iter)

    if options.infer_types:
        nulls = frozenset(options.null_values)
        string_cols = frozenset(options.string_columns)
        date_fmts = tuple(options.date_formats)
        dt_fmts = tuple(options.datetime_formats)
        keep_string_flags = [c in string_cols for c in columns]

        def _row_to_dict(raw_row):
            out: dict = {}
            for col, flag, value in zip(columns, keep_string_flags, raw_row):
                out[col] = _coerce_cell(
                    value,
                    null_values=nulls,
                    thousands_sep=options.thousands_separator,
                    decimal_sep=options.decimal_separator,
                    date_formats=date_fmts,
                    datetime_formats=dt_fmts,
                    keep_string=flag,
                )
            return out
    else:
        def _row_to_dict(raw_row):
            return dict(zip(columns, raw_row))

    row_size = options.row_size if options.row_size else 4096
    rows: list[dict] = []
    for raw_row in row_iter:
        rows.append(_row_to_dict(raw_row))
        if len(rows) >= row_size:
            yield from _rows_to_batches(rows)
            rows = []
    if rows:
        yield from _rows_to_batches(rows)


def _write_sheet_rows(
    ws: Any,
    batches: Iterable[pa.RecordBatch],
    *,
    has_header: bool,
) -> None:
    """Append batches to *ws* in row order, optionally writing a header.

    The first batch's schema names are taken as the column order;
    subsequent batches re-use that order via ``row.get(name)`` so
    schema drift across batches doesn't crash the writer.
    """
    iterator = iter(batches)
    first = next(iterator, None)
    if first is None:
        return
    column_names = first.schema.names
    if has_header:
        ws.append(column_names)
    for row in first.to_pylist():
        ws.append([row.get(c) for c in column_names])
    for batch in iterator:
        for row in batch.to_pylist():
            ws.append([row.get(c) for c in column_names])


# ---------------------------------------------------------------------------
# XlsxSheetIO — lazy per-sheet child
# ---------------------------------------------------------------------------


class XlsxSheetIO(BytesIO):
    """:class:`BytesIO` over a single worksheet's rows.

    A sheet has no standalone byte representation — the workbook
    keeps everything in one ZIP archive. The Tabular hooks read /
    write the sheet directly through openpyxl on the parent
    workbook; the byte surface (``to_bytes`` / ``read``) materializes
    a CSV view of the sheet on demand for callers that want one,
    matching the lazy-byte-surface contract :class:`ZipEntryIO` set.

    Tabular dispatch picks this class up via the ``XLSX_SHEET``
    mime — but the registry lookup is rarely needed: the parent
    :class:`XlsxIO` constructs sheets directly through
    :meth:`XlsxIO.iter_children` / :meth:`XlsxIO.child`.
    """

    __slots__ = (
        "sheet_name",
        "_xlsx_parent",
        "_materialized",
    )

    def __init__(
        self,
        *,
        sheet_name: str,
        xlsx_parent: "XlsxIO",
        **kwargs: Any,
    ) -> None:
        super().__init__(holder=Memory(), owns_holder=True, **kwargs)
        self.sheet_name: str = sheet_name
        self._xlsx_parent: "XlsxIO" = xlsx_parent
        self._materialized: bool = False

    @classmethod
    def options_class(cls):
        return XlsxOptions

    # ==================================================================
    # Lazy materialization — render this sheet as CSV bytes on demand
    # ==================================================================

    def _materialize(self) -> None:
        if self._materialized:
            return
        if self._xlsx_parent.size == 0:
            self._materialized = True
            return

        # Read the sheet via openpyxl, encode as CSV, stash in self's
        # Memory holder. The CSV is purely a convenience for callers
        # that drop to byte-level — Tabular reads bypass it.
        import csv as _csv
        openpyxl = _openpyxl()
        with self._xlsx_parent._format_view() as v:
            v.seek(0)
            wb = openpyxl.load_workbook(v, read_only=True, data_only=True)
            try:
                if self.sheet_name not in wb.sheetnames:
                    self._materialized = True
                    return
                ws = wb[self.sheet_name]
                buf = _stdlib_io.StringIO(newline="")
                writer = _csv.writer(buf, lineterminator="\n")
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(["" if v is None else v for v in row])
                payload = buf.getvalue().encode("utf-8")
            finally:
                wb.close()
        self._holder.write_bytes(payload, 0)
        self._materialized = True

    def _active(self):
        if not self._materialized:
            self._materialize()
        return super()._active()

    @property
    def size(self) -> int:
        if not self._materialized:
            self._materialize()
        return self._holder.size

    # ==================================================================
    # Tabular hooks — straight through openpyxl on the parent
    # ==================================================================

    def _collect_schema(self, options: XlsxOptions) -> Schema:
        first = next(iter(self._read_arrow_batches(options)), None)
        if first is None:
            return Schema.empty()
        return Schema.from_arrow(first.schema)

    def _read_arrow_batches(
        self,
        options: XlsxOptions,
    ) -> Iterator[pa.RecordBatch]:
        if self._xlsx_parent.size == 0:
            return
        openpyxl = _openpyxl()
        with self._xlsx_parent._format_view() as v:
            v.seek(0)
            wb = openpyxl.load_workbook(v, read_only=True, data_only=True)
            try:
                if self.sheet_name not in wb.sheetnames:
                    return
                ws = wb[self.sheet_name]
                yield from _stream_sheet_batches(ws, options=options)
            finally:
                wb.close()

    def _write_arrow_batches(
        self,
        batches: Iterable[pa.RecordBatch],
        options: XlsxOptions,
    ) -> None:
        """Update this sheet in the parent workbook.

        Read-modify-rewrite the parent: copy every other sheet
        through (cell by cell), drop the existing sheet with this
        name (if any), then append the new rows under
        ``self.sheet_name``. The whole thing rewrites the workbook
        bytes — xlsx has no incremental update story at the leaf
        level.
        """
        target_name = self.sheet_name
        action = options.mode
        if action is Mode.IGNORE and self._xlsx_parent.size > 0:
            return
        if action is Mode.ERROR_IF_EXISTS and self._xlsx_parent.size > 0:
            raise FileExistsError(
                f"{type(self).__name__} parent workbook is non-empty; "
                f"refusing to overwrite under mode={options.mode!r}."
            )

        # Snapshot the parent's other sheets so we can re-emit them
        # alongside the new one. Empty parent → just write a fresh
        # workbook with one sheet.
        carry: list[tuple[str, list[tuple]]] = []
        if self._xlsx_parent.size > 0:
            openpyxl = _openpyxl()
            with self._xlsx_parent._format_view() as v:
                v.seek(0)
                wb = openpyxl.load_workbook(v, read_only=True, data_only=True)
                try:
                    for name in wb.sheetnames:
                        if name == target_name:
                            continue
                        ws = wb[name]
                        carry.append(
                            (name, [tuple(row) for row in ws.iter_rows(values_only=True)])
                        )
                finally:
                    wb.close()

        # Build the new workbook in a streaming write.
        openpyxl = _openpyxl()
        wb_out = openpyxl.Workbook(write_only=True)
        try:
            for name, rows in carry:
                ws = wb_out.create_sheet(title=name)
                for row in rows:
                    ws.append(list(row))
            ws_target = wb_out.create_sheet(title=target_name)
            _write_sheet_rows(
                ws_target, batches, has_header=options.has_header,
            )

            scratch = _stdlib_io.BytesIO()
            wb_out.save(scratch)
            # Commit through the parent so its codec layer is honored.
            self._xlsx_parent._commit_format_payload(scratch.getbuffer())
        finally:
            wb_out.close()

        # Invalidate our cached materialization — the sheet's bytes
        # may have changed.
        self._materialized = False
        self._holder.write_bytes(b"", 0)

    def __repr__(self) -> str:
        state = "materialized" if self._materialized else "lazy"
        return (
            f"<{type(self).__name__} {self.sheet_name!r} "
            f"{state} parent={self._xlsx_parent!r}>"
        )


# ---------------------------------------------------------------------------
# XlsxIO — workbook-level surface
# ---------------------------------------------------------------------------


class XlsxIO(BytesIO):
    """:class:`Tabular` leaf for xlsx workbooks (single- or multi-sheet)."""

    mime_type: ClassVar[MimeTypes] = MimeTypes.XLSX

    @classmethod
    def options_class(cls):
        return XlsxOptions

    # ==================================================================
    # Children surface — lazy per-sheet iteration
    # ==================================================================

    def list_sheets(self) -> "list[str]":
        """Return sheet names in workbook order. One openpyxl pass; no row reads."""
        if self.size == 0:
            return []
        openpyxl = _openpyxl()
        with self._format_view() as v:
            v.seek(0)
            wb = openpyxl.load_workbook(v, read_only=True, data_only=True)
            try:
                return list(wb.sheetnames)
            finally:
                wb.close()

    def iter_children(self) -> Iterator[XlsxSheetIO]:
        """Yield every sheet as a lazy :class:`XlsxSheetIO`.

        The directory walk is one ``load_workbook`` call; per-sheet
        rows are NOT pulled until the caller hits the child's
        Tabular hook (:meth:`read_arrow_table`,
        :meth:`collect_schema`, …) or a byte-level op.
        """
        for name in self.list_sheets():
            yield self.adopt_child(
                XlsxSheetIO(sheet_name=name, xlsx_parent=self)
            )

    def child(self, sheet_name: str) -> XlsxSheetIO:
        """Return a lazy :class:`XlsxSheetIO` for *sheet_name*.

        Raises :class:`KeyError` when the workbook doesn't contain
        a sheet with that name. Fetching the directory is the only
        eager work — sheet rows materialize on first read.
        """
        sheets = self.list_sheets()
        if sheet_name not in sheets:
            raise KeyError(
                f"No sheet named {sheet_name!r} in {self!r}. "
                f"Available: {sheets!r}."
            )
        return self.adopt_child(
            XlsxSheetIO(sheet_name=sheet_name, xlsx_parent=self)
        )

    # ==================================================================
    # Schema
    # ==================================================================

    def _collect_schema(self, options: XlsxOptions) -> Schema:
        if self.size == 0:
            return Schema.empty()
        first = next(iter(self._read_arrow_batches(options)), None)
        if first is None:
            return Schema.empty()
        return Schema.from_arrow(first.schema)

    # ==================================================================
    # Read path — single sheet, picked by ``options.sheet_name``
    # ==================================================================

    def _read_arrow_batches(
        self,
        options: XlsxOptions,
    ) -> Iterator[pa.RecordBatch]:
        """Stream rows from the named sheet, batch them, yield.

        For multi-sheet walks use :meth:`iter_children` and call
        :meth:`XlsxSheetIO.read_arrow_batches` per sheet — concatenating
        sheets with different shapes at this level would silently
        corrupt the schema.
        """
        if self.size == 0:
            return

        openpyxl = _openpyxl()
        # ``_format_view`` peels any codec on the buffer's MediaType
        # (e.g. ``xlsx + gzip`` from a ``.xlsx.gz`` LocalPath) so
        # openpyxl receives the uncompressed ZIP bytes.
        with self._format_view() as v:
            v.seek(0)
            wb = openpyxl.load_workbook(v, read_only=True, data_only=True)
            try:
                ws = (
                    wb[options.sheet_name]
                    if options.sheet_name in wb.sheetnames
                    else wb[wb.sheetnames[0]]
                )
                yield from _stream_sheet_batches(ws, options=options)
            finally:
                wb.close()

    # ==================================================================
    # Write path
    # ==================================================================

    def _write_arrow_batches(
        self,
        batches: Iterable[pa.RecordBatch],
        options: XlsxOptions,
    ) -> None:
        """Persist *batches* as a single sheet in the workbook.

        Mode dispatch:

        - **OVERWRITE / AUTO / TRUNCATE** — fresh workbook with one
          sheet.
        - **APPEND** / **UPSERT** / **MERGE** — preserve every other
          sheet in the existing workbook, replace
          ``options.sheet_name`` with the incoming batches, and
          rewrite the archive.
        - **IGNORE** — skip when non-empty.
        - **ERROR_IF_EXISTS** — raise when non-empty.
        """
        action = self._resolve_action(options.mode)

        if action is Mode.IGNORE:
            if self.size > 0:
                return
            action = Mode.OVERWRITE
        elif action is Mode.ERROR_IF_EXISTS:
            if self.size > 0:
                raise FileExistsError(
                    f"{type(self).__name__} buffer is non-empty "
                    f"({self.size} bytes); refusing to overwrite under "
                    f"mode={options.mode!r}."
                )
            action = Mode.OVERWRITE

        iterator = iter(batches)
        first = next(iterator, None)
        if first is None and action is Mode.OVERWRITE:
            self.seek(0)
            self.truncate(0)
            return
        if first is None:
            return

        # Pull the existing workbook's other sheets through so we
        # can re-emit them alongside the new one (APPEND only).
        carry: list[tuple[str, list[tuple]]] = []
        if action is Mode.APPEND and self.size > 0:
            openpyxl = _openpyxl()
            with self._format_view() as v:
                v.seek(0)
                wb_in = openpyxl.load_workbook(v, read_only=True, data_only=True)
                try:
                    for name in wb_in.sheetnames:
                        if name == options.sheet_name:
                            continue
                        ws = wb_in[name]
                        carry.append(
                            (name, [tuple(row) for row in ws.iter_rows(values_only=True)])
                        )
                finally:
                    wb_in.close()

        openpyxl = _openpyxl()
        wb = openpyxl.Workbook(write_only=True)
        try:
            for name, rows in carry:
                ws = wb.create_sheet(title=name)
                for row in rows:
                    ws.append(list(row))
            ws_target = wb.create_sheet(title=options.sheet_name)
            _write_sheet_rows(
                ws_target,
                _chain_one(first, iterator),
                has_header=options.has_header,
            )
            # Save into an in-memory scratch buffer, then bulk-commit
            # via ``_commit_format_payload`` so a codec on the buffer's
            # MediaType (e.g. ``.xlsx.gz``) is applied transparently.
            scratch = _stdlib_io.BytesIO()
            wb.save(scratch)
            self._commit_format_payload(scratch.getbuffer())
        finally:
            wb.close()

    def write_sheets(
        self,
        sheets: "Mapping[str, pa.Table | pa.RecordBatch | Iterable[pa.RecordBatch]]",
        *,
        has_header: bool = True,
    ) -> None:
        """Pack ``{name: arrow_table}`` into a fresh workbook.

        Convenience for the common multi-sheet write — equivalent to
        looping :meth:`XlsxSheetIO._write_arrow_batches` once per
        sheet but emits the whole workbook in a single openpyxl
        write-only pass.
        """
        if not sheets:
            self.seek(0)
            self.truncate(0)
            return

        openpyxl = _openpyxl()
        wb = openpyxl.Workbook(write_only=True)
        try:
            for name, payload in sheets.items():
                ws = wb.create_sheet(title=name)
                if isinstance(payload, pa.Table):
                    batches = payload.to_batches()
                elif isinstance(payload, pa.RecordBatch):
                    batches = [payload]
                else:
                    batches = list(payload)
                _write_sheet_rows(ws, batches, has_header=has_header)

            scratch = _stdlib_io.BytesIO()
            wb.save(scratch)
            self._commit_format_payload(scratch.getbuffer())
        finally:
            wb.close()

    def _resolve_action(self, mode: Mode) -> Mode:
        if mode is Mode.AUTO or mode is Mode.OVERWRITE or mode is Mode.TRUNCATE:
            return Mode.OVERWRITE
        if mode is Mode.IGNORE:
            return Mode.IGNORE
        if mode is Mode.ERROR_IF_EXISTS:
            return Mode.ERROR_IF_EXISTS
        if mode is Mode.APPEND or mode is Mode.UPSERT or mode is Mode.MERGE:
            return Mode.APPEND
        return Mode.OVERWRITE
