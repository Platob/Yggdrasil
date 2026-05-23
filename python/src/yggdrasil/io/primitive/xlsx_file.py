"""XLSX Tabular leaf with lazy per-sheet entries.

:class:`XLSXFile` IS-A :class:`BytesIO` whose backing bytes are an
xlsx workbook (a ZIP archive). Mirrors the :class:`ZipFile` shape:

1. **Byte surface** — inherited from :class:`BytesIO`. Read / write
   the raw workbook bytes.
2. **Children surface** — :meth:`iter_children` walks every
   worksheet as a :class:`XLSXSheetFile`. Sheets are **lazy**: rows
   are pulled out of the parent workbook on first read.

Reads use :func:`fastexcel.read_excel` (calamine-based, returns
arrow record batches directly). Writes use
:func:`openpyxl.Workbook(write_only=True)` since fastexcel is
read-only and we need the multi-sheet APPEND story.

Mode dispatch on a workbook write:

- **OVERWRITE / AUTO / TRUNCATE** — fresh workbook with one sheet.
- **APPEND / UPSERT / MERGE** — preserve every other sheet, replace
  ``options.sheet_name`` with the incoming batches.
- **IGNORE** — skip when non-empty.
- **ERROR_IF_EXISTS** — raise when non-empty.

Convenience helper :meth:`write_sheets` packs a ``{name: arrow_table}``
mapping into a fresh workbook.
"""

from __future__ import annotations

import dataclasses
import io as _stdlib_io
import itertools as _it
from typing import Any, ClassVar, Iterable, Iterator, Mapping

import pyarrow as pa

from yggdrasil.data.options import CastOptions
from yggdrasil.data.schema import Schema
from yggdrasil.data.enums import MimeTypes, Mode
from yggdrasil.io.base import IO
from yggdrasil.io.memory import Memory

__all__ = ["XLSXFile", "XlsxOptions", "XLSXSheetFile"]


@dataclasses.dataclass(frozen=True, slots=True)
class XlsxOptions(CastOptions):
    """:class:`CastOptions` extended with XLSX-specific knobs."""

    #: Active worksheet for single-sheet reads / writes. On read,
    #: missing names fall back to the first sheet; on write, this is
    #: the sheet that receives the incoming batches.
    sheet_name: str = "Sheet1"
    has_header: bool = True


def _openpyxl():
    """Lazy openpyxl import — used only on the write path."""
    try:
        import openpyxl
    except ImportError as e:  # pragma: no cover
        raise ImportError(
            "XLSXFile writes require openpyxl. Install with: pip install openpyxl"
        ) from e
    return openpyxl


def _fastexcel():
    """Lazy fastexcel import — used for sheet listing on the read path."""
    try:
        import fastexcel
    except ImportError as e:  # pragma: no cover
        raise ImportError(
            "XLSXFile reads require fastexcel. Install with: "
            "pip install 'ygg[excel]'"
        ) from e
    return fastexcel


def _list_sheet_names(data: bytes) -> "list[str]":
    if not data:
        return []
    return list(_fastexcel().read_excel(data).sheet_names)


def _read_sheet_batch(
    data: bytes,
    *,
    sheet_name: str,
    has_header: bool,
    fallback_to_first: bool,
) -> "pa.RecordBatch | None":
    """Read a single sheet via fastexcel into a pyarrow RecordBatch.

    Returns ``None`` when *data* is empty or the sheet doesn't exist
    (and *fallback_to_first* is False).
    """
    if not data:
        return None
    parser = _fastexcel().read_excel(data)
    sheets = parser.sheet_names
    if not sheets:
        return None
    if sheet_name not in sheets:
        if not fallback_to_first:
            return None
        sheet_name = sheets[0]
    return parser.load_sheet(
        sheet_name,
        header_row=0 if has_header else None,
        eager=True,
    )


def _write_sheet_rows(
    ws: Any,
    batches: Iterable[pa.RecordBatch],
    *,
    has_header: bool,
) -> None:
    """Append batches to *ws*, optionally writing a header.

    The first batch's schema names set column order; subsequent
    batches re-use that order via ``row.get(name)`` so schema drift
    across batches doesn't crash the writer.
    """
    iterator = iter(batches)
    first = next(iterator, None)
    if first is None:
        return
    column_names = first.schema.names
    if has_header:
        ws.append(column_names)
    for row in first.to_pylist():
        ws.append([row.get(c) for c in column_names])
    for batch in iterator:
        for row in batch.to_pylist():
            ws.append([row.get(c) for c in column_names])


# ---------------------------------------------------------------------------
# XLSXSheetFile — lazy per-sheet child
# ---------------------------------------------------------------------------


class XLSXSheetFile(IO[bytes, XlsxOptions]):
    """:class:`BytesIO` over a single worksheet's rows.

    A sheet has no standalone byte representation — the workbook
    keeps everything in one ZIP archive. The Tabular hooks read /
    write the sheet directly through the parent workbook; the byte
    surface (``to_bytes`` / ``read``) materializes a CSV view of the
    sheet on demand for callers that drop to byte-level.
    """

    __slots__ = (
        "sheet_name",
        "_xlsx_parent",
        "_materialized",
    )

    def __init__(
        self,
        *,
        sheet_name: str,
        xlsx_parent: "XLSXFile",
        **kwargs: Any,
    ) -> None:
        super().__init__(holder=Memory(), owns_holder=True, **kwargs)
        self.sheet_name: str = sheet_name
        self._xlsx_parent: "XLSXFile" = xlsx_parent
        self._materialized: bool = False

    @classmethod
    def options_class(cls):
        return XlsxOptions

    # ==================================================================
    # Lazy materialization — render this sheet as CSV bytes on demand
    # ==================================================================

    def _materialize(self) -> None:
        if self._materialized:
            return
        if self._xlsx_parent.size == 0:
            self._materialized = True
            return

        with self._xlsx_parent.arrow_input_stream() as v:
            data = v.read()
        batch = _read_sheet_batch(
            data,
            sheet_name=self.sheet_name,
            has_header=True,
            fallback_to_first=False,
        )
        if batch is None:
            self._materialized = True
            return
        from pyarrow import csv as _pa_csv
        sink = pa.BufferOutputStream()
        _pa_csv.write_csv(
            pa.Table.from_batches([batch]),
            sink,
            write_options=_pa_csv.WriteOptions(
                quoting_style="none", quoting_header="none",
            ),
        )
        self._parent.write_bytes(sink.getvalue(), 0)
        self._materialized = True

    def _active(self):
        if not self._materialized:
            self._materialize()
        return super()._active()

    @property
    def size(self) -> int:
        if not self._materialized:
            self._materialize()
        return self._parent.size

    # ==================================================================
    # Tabular hooks
    # ==================================================================

    def _collect_schema(self, options: XlsxOptions) -> Schema:
        first = next(iter(self._read_arrow_batches(options)), None)
        if first is None:
            return Schema.empty()
        return Schema.from_arrow(first.schema)

    def _read_arrow_batches(
        self,
        options: XlsxOptions,
    ) -> Iterator[pa.RecordBatch]:
        if self._xlsx_parent.size == 0:
            return
        with self._xlsx_parent.arrow_input_stream() as v:
            data = v.read()
        batch = _read_sheet_batch(
            data,
            sheet_name=self.sheet_name,
            has_header=options.has_header,
            fallback_to_first=False,
        )
        if batch is not None:
            yield batch

    def _write_arrow_batches(
        self,
        batches: Iterable[pa.RecordBatch],
        options: XlsxOptions,
    ) -> None:
        """Update this sheet in the parent workbook.

        Read-modify-rewrite the parent: copy every other sheet
        through, drop the existing sheet with this name (if any),
        then append the new rows under ``self.sheet_name``. The
        whole thing rewrites the workbook bytes — xlsx has no
        incremental update story at the leaf level.
        """
        target_name = self.sheet_name
        action = options.mode
        if action is Mode.IGNORE and self._xlsx_parent.size > 0:
            return
        if action is Mode.ERROR_IF_EXISTS and self._xlsx_parent.size > 0:
            raise FileExistsError(
                f"{type(self).__name__} parent workbook is non-empty; "
                f"refusing to overwrite under mode={options.mode!r}."
            )

        carry: list[tuple[str, list[tuple]]] = []
        if self._xlsx_parent.size > 0:
            openpyxl = _openpyxl()
            with self._xlsx_parent.arrow_input_stream() as v:
                wb = openpyxl.load_workbook(v, read_only=True, data_only=True)
                try:
                    for name in wb.sheetnames:
                        if name == target_name:
                            continue
                        ws = wb[name]
                        carry.append(
                            (name, [tuple(row) for row in ws.iter_rows(values_only=True)])
                        )
                finally:
                    wb.close()

        openpyxl = _openpyxl()
        wb_out = openpyxl.Workbook(write_only=True)
        try:
            for name, rows in carry:
                ws = wb_out.create_sheet(title=name)
                for row in rows:
                    ws.append(list(row))
            ws_target = wb_out.create_sheet(title=target_name)
            _write_sheet_rows(
                ws_target, batches, has_header=options.has_header,
            )

            # openpyxl writes a ZIP archive and needs a seekable sink;
            # ``pa.BufferOutputStream`` is forward-only, so save into a
            # stdlib ``BytesIO`` scratch and bulk-commit through the
            # IO's :meth:`arrow_output_stream` (which applies any codec
            # and lays the bytes down on the durable holder on exit).
            scratch = _stdlib_io.BytesIO()
            wb_out.save(scratch)
            with self._xlsx_parent.arrow_output_stream() as sink:
                sink.write(scratch.getvalue())
        finally:
            wb_out.close()

        self._materialized = False
        self._parent.write_bytes(b"", 0)

    def __repr__(self) -> str:
        state = "materialized" if self._materialized else "lazy"
        return (
            f"<{type(self).__name__} {self.sheet_name!r} "
            f"{state} parent={self._xlsx_parent!r}>"
        )


# ---------------------------------------------------------------------------
# XLSXFile — workbook-level surface
# ---------------------------------------------------------------------------


class XLSXFile(IO[bytes, XlsxOptions]):
    """:class:`Tabular` leaf for xlsx workbooks (single- or multi-sheet)."""

    mime_type: ClassVar[MimeTypes] = MimeTypes.XLSX

    @classmethod
    def options_class(cls):
        return XlsxOptions

    # ==================================================================
    # Children surface — lazy per-sheet iteration
    # ==================================================================

    def list_sheets(self) -> "list[str]":
        """Return sheet names in workbook order. One fastexcel pass; no row reads."""
        if self.size == 0:
            return []
        with self.arrow_input_stream() as v:
            return _list_sheet_names(v.read())

    def iter_children(self) -> Iterator[XLSXSheetFile]:
        """Yield every sheet as a lazy :class:`XLSXSheetFile`.

        The directory walk is one ``fastexcel.read_excel`` call;
        per-sheet rows are NOT pulled until the caller hits the
        child's Tabular hook (:meth:`read_arrow_table`,
        :meth:`collect_schema`, …) or a byte-level op.
        """
        for name in self.list_sheets():
            yield self.adopt_child(
                XLSXSheetFile(sheet_name=name, xlsx_parent=self)
            )

    def child(self, sheet_name: str) -> XLSXSheetFile:
        """Return a lazy :class:`XLSXSheetFile` for *sheet_name*.

        Raises :class:`KeyError` when the workbook doesn't contain
        a sheet with that name.
        """
        sheets = self.list_sheets()
        if sheet_name not in sheets:
            raise KeyError(
                f"No sheet named {sheet_name!r} in {self!r}. "
                f"Available: {sheets!r}."
            )
        return self.adopt_child(
            XLSXSheetFile(sheet_name=sheet_name, xlsx_parent=self)
        )

    # ==================================================================
    # Schema
    # ==================================================================

    def _collect_schema(self, options: XlsxOptions) -> Schema:
        if self.size == 0:
            return Schema.empty()
        first = next(iter(self._read_arrow_batches(options)), None)
        if first is None:
            return Schema.empty()
        return Schema.from_arrow(first.schema)

    # ==================================================================
    # Read path — single sheet, picked by ``options.sheet_name``
    # ==================================================================

    def _read_arrow_batches(
        self,
        options: XlsxOptions,
    ) -> Iterator[pa.RecordBatch]:
        """Stream rows from the named sheet, batch them, yield.

        For multi-sheet walks use :meth:`iter_children` and call
        :meth:`XLSXSheetFile.read_arrow_batches` per sheet — concatenating
        sheets with different shapes at this level would silently
        corrupt the schema.
        """
        if self.size == 0:
            return
        # ``arrow_input_stream`` peels any codec on the buffer's
        # MediaType (e.g. ``xlsx + gzip`` from a ``.xlsx.gz``
        # LocalPath) so the parser receives the uncompressed ZIP
        # bytes — and uses :func:`pyarrow.memory_map` for local-path
        # holders so the workbook lands in the page cache once.
        with self.arrow_input_stream() as v:
            data = v.read()
        batch = _read_sheet_batch(
            data,
            sheet_name=options.sheet_name,
            has_header=options.has_header,
            fallback_to_first=True,
        )
        if batch is not None:
            yield batch

    # ==================================================================
    # Write path
    # ==================================================================

    def _write_arrow_batches(
        self,
        batches: Iterable[pa.RecordBatch],
        options: XlsxOptions,
    ) -> None:
        """Persist *batches* as a single sheet in the workbook.

        Mode dispatch:

        - **OVERWRITE / AUTO / TRUNCATE** — fresh workbook with one
          sheet.
        - **APPEND** / **UPSERT** / **MERGE** — preserve every other
          sheet in the existing workbook, replace
          ``options.sheet_name`` with the incoming batches, and
          rewrite the archive.
        - **IGNORE** — skip when non-empty.
        - **ERROR_IF_EXISTS** — raise when non-empty.
        """
        action = self._resolve_action(options.mode)

        _has_existing = self.size_known and self.size > 0
        if action is Mode.IGNORE:
            if _has_existing:
                return
            action = Mode.OVERWRITE
        elif action is Mode.ERROR_IF_EXISTS:
            if _has_existing:
                raise FileExistsError(
                    f"{type(self).__name__} buffer is non-empty "
                    f"({self.size} bytes); refusing to overwrite under "
                    f"mode={options.mode!r}."
                )
            action = Mode.OVERWRITE

        iterator = iter(batches)
        first = next(iterator, None)
        if first is None and action is Mode.OVERWRITE:
            self.seek(0)
            self.truncate(0)
            return
        if first is None:
            return

        carry: list[tuple[str, list[tuple]]] = []
        if action is Mode.APPEND and _has_existing:
            openpyxl = _openpyxl()
            with self.arrow_input_stream() as v:
                wb_in = openpyxl.load_workbook(v, read_only=True, data_only=True)
                try:
                    for name in wb_in.sheetnames:
                        if name == options.sheet_name:
                            continue
                        ws = wb_in[name]
                        carry.append(
                            (name, [tuple(row) for row in ws.iter_rows(values_only=True)])
                        )
                finally:
                    wb_in.close()

        openpyxl = _openpyxl()
        wb = openpyxl.Workbook(write_only=True)
        try:
            for name, rows in carry:
                ws = wb.create_sheet(title=name)
                for row in rows:
                    ws.append(list(row))
            ws_target = wb.create_sheet(title=options.sheet_name)
            _write_sheet_rows(
                ws_target,
                _it.chain([first], iterator),
                has_header=options.has_header,
            )
            # openpyxl writes a ZIP archive and needs a seekable sink;
            # ``pa.BufferOutputStream`` is forward-only, so save into
            # a stdlib ``BytesIO`` scratch and bulk-commit through the
            # IO's :meth:`arrow_output_stream` — that applies any codec
            # on the buffer's MediaType (e.g. ``.xlsx.gz``) transparently.
            scratch = _stdlib_io.BytesIO()
            wb.save(scratch)
            with self.arrow_output_stream() as sink:
                sink.write(scratch.getvalue())
        finally:
            wb.close()

    def write_sheets(
        self,
        sheets: "Mapping[str, pa.Table | pa.RecordBatch | Iterable[pa.RecordBatch]]",
        *,
        has_header: bool = True,
    ) -> None:
        """Pack ``{name: arrow_table}`` into a fresh workbook.

        Convenience for the common multi-sheet write — equivalent to
        looping :meth:`XLSXSheetFile._write_arrow_batches` once per
        sheet but emits the whole workbook in a single openpyxl
        write-only pass.
        """
        if not sheets:
            self.seek(0)
            self.truncate(0)
            return

        openpyxl = _openpyxl()
        wb = openpyxl.Workbook(write_only=True)
        try:
            for name, payload in sheets.items():
                ws = wb.create_sheet(title=name)
                if isinstance(payload, pa.Table):
                    batches = payload.to_batches()
                elif isinstance(payload, pa.RecordBatch):
                    batches = [payload]
                else:
                    batches = list(payload)
                _write_sheet_rows(ws, batches, has_header=has_header)

            scratch = _stdlib_io.BytesIO()
            wb.save(scratch)
            with self.arrow_output_stream() as sink:
                sink.write(scratch.getvalue())
        finally:
            wb.close()

    def _resolve_action(self, mode: Mode) -> Mode:
        if mode is Mode.AUTO or mode is Mode.OVERWRITE or mode is Mode.TRUNCATE:
            return Mode.OVERWRITE
        if mode is Mode.IGNORE:
            return Mode.IGNORE
        if mode is Mode.ERROR_IF_EXISTS:
            return Mode.ERROR_IF_EXISTS
        if mode is Mode.APPEND or mode is Mode.UPSERT or mode is Mode.MERGE:
            return Mode.APPEND
        return Mode.OVERWRITE
