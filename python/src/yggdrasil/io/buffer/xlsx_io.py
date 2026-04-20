"""XLSX I/O on top of :class:`~yggdrasil.io.buffer.BytesIO`.

Engines
-------
* ``"fallback"`` — uses :mod:`openpyxl` for both read and write. Always
  available in environments where yggdrasil is installed (openpyxl is
  a soft dep). Multi-sheet read + write. No optional native deps.
* ``"polars"`` — uses :func:`polars.read_excel` and
  :meth:`polars.DataFrame.write_excel`. Fastest on large sheets.
  Requires :mod:`polars` (+ :mod:`xlsxwriter` for write,
  :mod:`openpyxl` or :mod:`fastexcel` for read).
* ``"pandas"`` — uses :func:`pandas.read_excel` and
  :meth:`pandas.DataFrame.to_excel`. Requires :mod:`pandas` +
  :mod:`openpyxl`.
* ``"auto"`` — prefers ``polars`` when available, falls back to the
  built-in openpyxl path otherwise. ``pandas`` is never chosen by
  ``auto`` — ask for it explicitly.

Multi-sheet
-----------
Each call to ``write_arrow_table`` writes or replaces exactly **one**
sheet identified by ``write_sheet_name``; other sheets already in the
workbook are preserved. This makes the natural idiom for multi-sheet
workbooks::

    io_ = MediaIO.make(buf, MimeTypes.XLSX)
    io_.write_arrow_table(t1, write_sheet_name="Summary",
                          mode=SaveMode.OVERWRITE)
    io_.write_arrow_table(t2, write_sheet_name="Details",
                          mode=SaveMode.APPEND)   # append preserves Summary

On read, ``sheet_name`` or ``sheet_id`` (0-based) selects the sheet;
defaults to the first sheet when both are ``None``.

Save modes
----------
Save modes refer to the **workbook as a whole**, not the individual
sheet. OVERWRITE truncates and starts a new workbook with the target
sheet. APPEND/UPSERT preserve other sheets and merge rows within the
target sheet (via read-old-then-rewrite; XLSX has no byte-level append).
IGNORE / ERROR_IF_EXISTS route through the base-class guard.

Transport-level compression (``MediaType.codec``) is handled by the
base class via ``open()`` / ``close()`` / ``mark_dirty()``.
"""
from __future__ import annotations

import io as _stdio
from dataclasses import dataclass
from typing import TYPE_CHECKING, Any, Iterator, Optional, Sequence

import pyarrow as pa

from yggdrasil.io.enums import SaveMode
from .media_io import MediaIO
from .media_options import MediaOptions

if TYPE_CHECKING:
    import pyarrow

__all__ = ["XlsxOptions", "XlsxIO"]


_VALID_ENGINES = frozenset({"auto", "fallback", "polars", "pandas"})
_DEFAULT_SHEET = "Sheet1"


def _has(module: str) -> bool:
    try:
        __import__(module)
        return True
    except ImportError:
        return False


# ---------------------------------------------------------------------
# Options
# ---------------------------------------------------------------------

@dataclass
class XlsxOptions(MediaOptions):
    """Options for XLSX I/O.

    Parameters
    ----------
    sheet_name:
        Read-side: name of the sheet to load. Ignored when ``sheet_id``
        is set. ``None`` means the first sheet.
    sheet_id:
        Read-side: 0-based index of the sheet to load. Overrides
        ``sheet_name`` when set. ``None`` means the first sheet.
    has_header:
        Read-side: whether the first (post-``skip_rows``) row is a
        header. When ``False``, columns are named ``f0, f1, f2, …``.
        Default ``True``.
    include_header:
        Write-side: emit column names as the first row. Default ``True``.
    skip_rows:
        Read-side: number of leading rows to skip *before* header
        detection. Default ``0``.
    write_sheet_name:
        Write-side: name of the sheet to write. If the buffer already
        contains other sheets (APPEND mode), the named sheet is
        replaced in place; other sheets are preserved. Default
        ``"Sheet1"``.
    engine:
        One of ``{"auto", "fallback", "polars", "pandas"}``. Default
        ``"auto"`` (prefers polars, falls back to openpyxl).
    """

    sheet_name: str | None = None
    sheet_id: int | None = None
    has_header: bool = True
    include_header: bool = True
    skip_rows: int = 0
    write_sheet_name: str = _DEFAULT_SHEET
    engine: str = "auto"

    def __post_init__(self) -> None:
        """Normalize and validate XLSX-specific options."""
        super().__post_init__()

        # sheet_name: str | None — reject bool (which is an int subclass
        # and would slip past an `isinstance(..., int)` check elsewhere).
        if self.sheet_name is not None:
            if isinstance(self.sheet_name, bool) or not isinstance(self.sheet_name, str):
                raise TypeError(
                    f"sheet_name must be str|None, got {type(self.sheet_name).__name__}"
                )
            if not self.sheet_name:
                raise ValueError("sheet_name must not be empty")

        # sheet_id: non-negative int.
        if self.sheet_id is not None:
            if isinstance(self.sheet_id, bool) or not isinstance(self.sheet_id, int):
                raise TypeError(
                    f"sheet_id must be int|None, got {type(self.sheet_id).__name__}"
                )
            if self.sheet_id < 0:
                raise ValueError(f"sheet_id must be non-negative, got {self.sheet_id}")

        if not isinstance(self.has_header, bool):
            raise TypeError(
                f"has_header must be bool, got {type(self.has_header).__name__}"
            )
        if not isinstance(self.include_header, bool):
            raise TypeError(
                f"include_header must be bool, got {type(self.include_header).__name__}"
            )

        if not isinstance(self.skip_rows, int) or isinstance(self.skip_rows, bool):
            raise TypeError(
                f"skip_rows must be int, got {type(self.skip_rows).__name__}"
            )
        if self.skip_rows < 0:
            raise ValueError(f"skip_rows must be non-negative, got {self.skip_rows}")

        if not isinstance(self.write_sheet_name, str):
            raise TypeError(
                f"write_sheet_name must be str, got {type(self.write_sheet_name).__name__}"
            )
        if not self.write_sheet_name:
            raise ValueError("write_sheet_name must not be empty")

        # Normalize engine case-insensitively, then validate.
        if not isinstance(self.engine, str):
            raise TypeError(
                f"engine must be str, got {type(self.engine).__name__}"
            )
        self.engine = self.engine.lower()
        if self.engine not in _VALID_ENGINES:
            raise ValueError(
                f"engine must be one of {sorted(_VALID_ENGINES)}, "
                f"got {self.engine!r}"
            )

    @classmethod
    def resolve(cls, *, options: "XlsxOptions | None" = None, **overrides: Any) -> "XlsxOptions":
        return cls.check_parameters(options=options, **overrides)


# ---------------------------------------------------------------------
# XlsxIO
# ---------------------------------------------------------------------

@dataclass(slots=True)
class XlsxIO(MediaIO[XlsxOptions]):
    """XLSX I/O with multi-engine support and multi-sheet semantics."""

    @classmethod
    def check_options(
        cls,
        options: Optional[XlsxOptions],
        *args,
        **kwargs,
    ) -> XlsxOptions:
        return XlsxOptions.check_parameters(options=options, **kwargs)

    # ------------------------------------------------------------------
    # Engine resolution
    # ------------------------------------------------------------------

    @staticmethod
    def _resolve_engine(requested: str) -> str:
        """Turn ``"auto"`` into a concrete engine name."""
        if requested != "auto":
            return requested
        if _has("polars"):
            return "polars"
        return "fallback"

    # ------------------------------------------------------------------
    # Core read protocol
    # ------------------------------------------------------------------

    def _read_arrow_batches(
        self,
        options: XlsxOptions,
    ) -> Iterator["pyarrow.RecordBatch"]:
        """Yield Arrow batches from the XLSX buffer."""
        with self.open() as b:
            if b.buffer.size <= 0:
                return

            engine = self._resolve_engine(options.engine)
            table = self._read_via_engine(engine, options)

            if table is None or table.num_rows == 0:
                if options.ignore_empty:
                    return
                # Still yield an empty batch with the schema so callers
                # can distinguish "no data" from "error" via iteration.
                if table is not None and len(table.schema) > 0:
                    yield from table.to_batches()
                return

            # Build a batch iterator, apply projection + cast.
            batch_size = getattr(options, "batch_size", 0) or 0
            if batch_size > 0:
                batches: Iterator[pa.RecordBatch] = iter(table.to_batches(max_chunksize=batch_size))
            else:
                batches = iter(table.to_batches())

            if options.columns is not None:
                schema_names = set(table.schema.names)
                wanted = [c for c in options.columns if c in schema_names]
                batches = (batch.select(wanted) for batch in batches)

            if options.ignore_empty:
                batches = (batch for batch in batches if batch.num_rows > 0)

            yield from options.cast.cast_iterator(batches)

    def _read_via_engine(
        self,
        engine: str,
        options: XlsxOptions,
    ) -> "pa.Table | None":
        """Dispatch to the concrete engine-specific reader."""
        if engine == "fallback":
            return self._read_fallback(options)
        if engine == "polars":
            return self._read_polars(options)
        if engine == "pandas":
            return self._read_pandas(options)
        raise ValueError(f"Unknown engine: {engine!r}")

    # ------------------------------------------------------------------
    # Fallback engine (openpyxl)
    # ------------------------------------------------------------------

    def _read_fallback(self, options: XlsxOptions) -> "pa.Table | None":
        """Read using openpyxl. Applies skip_rows, has_header, sheet select."""
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise ImportError(
                "The 'fallback' engine requires openpyxl. "
                "Install with: pip install openpyxl"
            ) from e

        data = self.buffer.to_bytes()
        if not data:
            return pa.table({})

        wb = load_workbook(
            _stdio.BytesIO(data),
            read_only=True,
            data_only=True,
        )
        try:
            ws = self._select_sheet(wb, options)
            rows_iter = ws.iter_rows(values_only=True)

            # Skip leading rows.
            for _ in range(options.skip_rows):
                try:
                    next(rows_iter)
                except StopIteration:
                    return pa.table({})

            # Materialize remaining rows. openpyxl's read_only mode
            # already streams from disk, but pa.Table construction needs
            # the full column-oriented data anyway.
            all_rows = [list(row) for row in rows_iter]
        finally:
            wb.close()

        return self._rows_to_table(all_rows, options.has_header)

    def _read_polars(self, options: XlsxOptions) -> "pa.Table | None":
        """Read via polars.read_excel."""
        try:
            import polars as pl
        except ImportError as e:
            raise ImportError(
                "The 'polars' engine requires polars. "
                "Install with: pip install polars"
            ) from e

        read_kwargs: dict[str, Any] = {
            "source": _stdio.BytesIO(self.buffer.to_bytes()),
            "has_header": options.has_header,
        }

        # Sheet selection. polars uses sheet_name (str) or sheet_id (int,
        # 1-based in older versions, 0-based in newer). Newer polars
        # also has a dedicated sheet_id parameter; fall back to sheet_name
        # via index lookup if sheet_id isn't directly supported.
        if options.sheet_name is not None:
            read_kwargs["sheet_name"] = options.sheet_name
        elif options.sheet_id is not None:
            # Polars' sheet_id is 1-based; our convention (and pandas') is
            # 0-based. Translate.
            read_kwargs["sheet_id"] = options.sheet_id + 1

        # skip_rows — polars' read_excel passes through read_options for
        # xlsx2csv; for the calamine backend, use the dedicated kwarg.
        if options.skip_rows > 0:
            # The exact kwarg path depends on polars version. Try the
            # modern flat kwarg first.
            read_kwargs["read_options"] = {"skip_rows": options.skip_rows}

        df = pl.read_excel(**read_kwargs)

        # polars may return a dict[str, DataFrame] when multiple sheets
        # are requested with no selection — our read_kwargs always
        # narrow to one sheet, so df is a DataFrame here.
        return df.to_arrow()

    def _read_pandas(self, options: XlsxOptions) -> "pa.Table | None":
        """Read via pandas.read_excel."""
        try:
            import pandas as pd
        except ImportError as e:
            raise ImportError(
                "The 'pandas' engine requires pandas. "
                "Install with: pip install pandas openpyxl"
            ) from e

        read_kwargs: dict[str, Any] = {
            "io": _stdio.BytesIO(self.buffer.to_bytes()),
            "header": 0 if options.has_header else None,
            "skiprows": options.skip_rows or None,
            "engine": "openpyxl",
        }

        if options.sheet_name is not None:
            read_kwargs["sheet_name"] = options.sheet_name
        elif options.sheet_id is not None:
            read_kwargs["sheet_name"] = options.sheet_id
        else:
            read_kwargs["sheet_name"] = 0

        df = pd.read_excel(**read_kwargs)

        # When has_header=False pandas gives integer column names;
        # normalize to f0, f1, ... to match the fallback/polars output.
        if not options.has_header:
            df.columns = [f"f{i}" for i in range(len(df.columns))]

        return pa.Table.from_pandas(df, preserve_index=False)

    # ------------------------------------------------------------------
    # Core write protocol
    # ------------------------------------------------------------------

    def _write_arrow_batches(
        self,
        batches: Iterator["pyarrow.RecordBatch"],
        options: XlsxOptions,
    ) -> None:
        """Write record batches as a single sheet into the XLSX buffer.

        Save-mode semantics apply to the *workbook*:
        - OVERWRITE / AUTO: replace the entire workbook.
        - APPEND / UPSERT: preserve other sheets; merge rows of the
          target sheet via read-old-then-rewrite.
        - IGNORE / ERROR_IF_EXISTS: handled by skip_write guard.
        """
        with self.open() as b:
            if self.skip_write(options.mode):
                return

            engine = self._resolve_engine(options.engine)

            # --- Collect new-side rows through the write cast --------
            cast_batches = options.cast.cast_arrow_tabular(batches)
            new_batches_list: list[pa.RecordBatch] = []
            for batch in cast_batches:
                if batch.num_rows == 0:
                    continue
                new_batches_list.append(batch)

            if not new_batches_list:
                new_table = pa.table({})
            else:
                new_table = pa.Table.from_batches(new_batches_list)

            # --- Save-mode: merge against existing target sheet ------
            # OVERWRITE/AUTO with a non-empty buffer means replace the
            # whole workbook. APPEND/UPSERT means preserve other sheets
            # and merge rows within the named sheet.
            preserve_other_sheets = (
                options.mode in (SaveMode.APPEND, SaveMode.UPSERT)
                and b.buffer.size > 0
            )

            merged_table = new_table
            other_sheets: dict[str, pa.Table] = {}

            if preserve_other_sheets:
                # Load existing workbook to get (a) the target sheet's
                # existing rows to merge with, (b) other sheets to keep.
                existing_sheets = self._load_all_sheets(options)
                target = existing_sheets.pop(options.write_sheet_name, None)
                other_sheets = existing_sheets

                if target is not None and target.num_rows > 0:
                    merged_table = self._merge_tables(
                        old=target,
                        new=new_table,
                        mode=options.mode,
                        match_by=options.match_by,
                    )

            # --- Truncate on OVERWRITE --------------------------------
            if not preserve_other_sheets and b.buffer.size > 0:
                b.buffer.truncate(0)

            # --- Write out the workbook -------------------------------
            self._write_workbook(
                target_sheet=options.write_sheet_name,
                target_table=merged_table,
                other_sheets=other_sheets,
                options=options,
                engine=engine,
            )
            b.mark_dirty()

    # ------------------------------------------------------------------
    # Multi-sheet helpers
    # ------------------------------------------------------------------

    def _load_all_sheets(self, options: XlsxOptions) -> dict[str, pa.Table]:
        """Load every sheet in the current workbook as an Arrow table.

        Always uses the fallback (openpyxl) engine since it's the only
        one that cleanly supports "give me all sheets as a dict" across
        arbitrary structures. Engine-specific read-side quirks don't
        apply here — this is internal to the APPEND/UPSERT rewrite.
        """
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise ImportError(
                "Multi-sheet APPEND/UPSERT requires openpyxl. "
                "Install with: pip install openpyxl"
            ) from e

        data = self.buffer.to_bytes()
        if not data:
            return {}

        wb = load_workbook(
            _stdio.BytesIO(data),
            read_only=True,
            data_only=True,
        )
        result: dict[str, pa.Table] = {}
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_iter = ws.iter_rows(values_only=True)
                all_rows = [list(row) for row in rows_iter]
                result[sheet_name] = self._rows_to_table(
                    all_rows, has_header=options.has_header
                )
        finally:
            wb.close()

        return result

    def _write_workbook(
        self,
        *,
        target_sheet: str,
        target_table: "pa.Table",
        other_sheets: dict[str, pa.Table],
        options: XlsxOptions,
        engine: str,
    ) -> None:
        """Write a full workbook: target_sheet + all other_sheets."""
        # When there's only one sheet to write and no preservation needed,
        # we can use the requested engine directly. Otherwise openpyxl
        # is the only engine that cleanly supports multi-sheet write.
        if not other_sheets and engine in ("polars", "pandas"):
            payload = self._write_single_sheet_via_engine(
                engine=engine,
                sheet_name=target_sheet,
                table=target_table,
                options=options,
            )
            self.buffer.replace_with_payload(payload)
            return

        # Fallback / multi-sheet path: openpyxl.
        payload = self._write_multi_sheet_fallback(
            target_sheet=target_sheet,
            target_table=target_table,
            other_sheets=other_sheets,
            options=options,
        )
        self.buffer.replace_with_payload(payload)

    def _write_single_sheet_via_engine(
        self,
        *,
        engine: str,
        sheet_name: str,
        table: "pa.Table",
        options: XlsxOptions,
    ) -> bytes:
        """Write *table* as a single sheet via polars or pandas."""
        if engine == "polars":
            try:
                import polars as pl
            except ImportError as e:
                raise ImportError(
                    "The 'polars' engine requires polars + xlsxwriter. "
                    "Install with: pip install polars xlsxwriter"
                ) from e

            df = pl.from_arrow(table)
            sink = _stdio.BytesIO()
            df.write_excel(
                workbook=sink,
                worksheet=sheet_name,
                include_header=options.include_header,
            )
            return sink.getvalue()

        if engine == "pandas":
            try:
                import pandas as pd
            except ImportError as e:
                raise ImportError(
                    "The 'pandas' engine requires pandas + openpyxl. "
                    "Install with: pip install pandas openpyxl"
                ) from e

            df = table.to_pandas()
            sink = _stdio.BytesIO()
            with pd.ExcelWriter(sink, engine="openpyxl") as writer:
                df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    header=options.include_header,
                    index=False,
                )
            return sink.getvalue()

        raise ValueError(f"Unsupported engine for single-sheet write: {engine!r}")

    def _write_multi_sheet_fallback(
        self,
        *,
        target_sheet: str,
        target_table: "pa.Table",
        other_sheets: dict[str, pa.Table],
        options: XlsxOptions,
    ) -> bytes:
        """Write a workbook containing target_sheet + preserved others."""
        try:
            from openpyxl import Workbook
        except ImportError as e:
            raise ImportError(
                "The 'fallback' engine requires openpyxl. "
                "Install with: pip install openpyxl"
            ) from e

        wb = Workbook(write_only=True)
        # Ordering: other sheets in their original order, then target.
        # Rationale: when the user calls write_arrow_table on a new
        # sheet name, they probably want it visible last (most recent).
        # If the target sheet already exists among other_sheets it was
        # popped out in _write_arrow_batches, so no dedup needed here.
        for name, tbl in other_sheets.items():
            ws = wb.create_sheet(title=self._safe_sheet_name(name))
            self._append_table_rows(ws, tbl, include_header=options.include_header)

        ws = wb.create_sheet(title=self._safe_sheet_name(target_sheet))
        self._append_table_rows(ws, target_table, include_header=options.include_header)

        sink = _stdio.BytesIO()
        wb.save(sink)
        return sink.getvalue()

    @staticmethod
    def _safe_sheet_name(name: str) -> str:
        """Sanitize a sheet name to meet XLSX's constraints.

        Excel forbids ``: \\ / ? * [ ]`` in sheet names and caps length
        at 31 chars. We replace forbidden chars with underscore and
        truncate.
        """
        forbidden = set(":\\/?*[]")
        cleaned = "".join("_" if c in forbidden else c for c in name)
        return cleaned[:31] or "Sheet"

    def _append_table_rows(
        self,
        ws: Any,
        table: "pa.Table",
        *,
        include_header: bool,
    ) -> None:
        """Append an Arrow table's rows to an openpyxl worksheet."""
        if include_header:
            ws.append(list(table.schema.names))

        if table.num_rows == 0:
            return

        # Iterate batch by batch to avoid materializing the full table
        # as Python objects up front on very large inputs.
        for batch in table.to_batches():
            for row in batch.to_pylist():
                # openpyxl expects ordered values; use the schema order.
                ws.append([row.get(col) for col in table.schema.names])

    # ------------------------------------------------------------------
    # Sheet selection + row → table conversion (fallback engine)
    # ------------------------------------------------------------------

    def _select_sheet(self, wb: Any, options: XlsxOptions) -> Any:
        """Resolve options to a concrete openpyxl worksheet."""
        if options.sheet_id is not None:
            sheets = wb.sheetnames
            if options.sheet_id >= len(sheets):
                raise IndexError(
                    f"sheet_id={options.sheet_id} out of range "
                    f"({len(sheets)} sheet(s) in workbook)"
                )
            return wb[sheets[options.sheet_id]]

        if options.sheet_name is not None:
            if options.sheet_name not in wb.sheetnames:
                raise KeyError(
                    f"Sheet {options.sheet_name!r} not in workbook "
                    f"(have: {wb.sheetnames})"
                )
            return wb[options.sheet_name]

        # Default: first sheet.
        return wb[wb.sheetnames[0]]

    @staticmethod
    def _rows_to_table(rows: list[list[Any]], has_header: bool) -> "pa.Table":
        """Convert a list-of-rows into an Arrow table.

        Column names come from the first row when *has_header* is True,
        otherwise use ``f0, f1, f2, …``. Empty input yields an empty
        table with no columns.
        """
        # Strip trailing empty rows that openpyxl sometimes appends for
        # worksheets with explicit dimensions.
        while rows and all(v is None for v in rows[-1]):
            rows.pop()

        if not rows:
            return pa.table({})

        if has_header:
            header = rows[0]
            body = rows[1:]
            column_names = [
                str(h) if h is not None else f"f{i}"
                for i, h in enumerate(header)
            ]
        else:
            body = rows
            ncols = max(len(row) for row in rows) if rows else 0
            column_names = [f"f{i}" for i in range(ncols)]

        if not column_names:
            return pa.table({})

        # Normalize row widths — openpyxl trims trailing None cells on
        # some sheets, which makes rows jagged.
        width = len(column_names)
        normalized = [
            list(row) + [None] * (width - len(row)) if len(row) < width else list(row)[:width]
            for row in body
        ]

        # Build column-oriented data.
        columns: dict[str, list[Any]] = {name: [] for name in column_names}
        for row in normalized:
            for name, value in zip(column_names, row):
                columns[name].append(value)

        return pa.table(columns)

    # ------------------------------------------------------------------
    # UPSERT / APPEND merge
    # ------------------------------------------------------------------

    @staticmethod
    def _merge_tables(
        *,
        old: "pa.Table",
        new: "pa.Table",
        mode: SaveMode,
        match_by: Any,
    ) -> "pa.Table":
        """Combine *old* and *new* according to the save mode.

        For APPEND: concatenate, unifying schemas (missing columns fill
        with null). For UPSERT: drop rows of *old* whose ``match_by``
        keys appear in *new*, then concatenate.
        """
        if mode == SaveMode.APPEND:
            return pa.concat_tables([old, new], promote_options="default")

        if mode == SaveMode.UPSERT:
            keys = _normalize_match_by(match_by)
            if not keys:
                raise ValueError("SaveMode.UPSERT requires options.match_by to be set")

            # Build a set of new-side key tuples.
            missing = [k for k in keys if k not in new.schema.names]
            if missing:
                raise KeyError(f"match_by columns not found in new table: {missing}")

            new_keys: set[tuple] = set()
            new_key_cols = [new.column(k).to_pylist() for k in keys]
            for i in range(new.num_rows):
                new_keys.add(tuple(col[i] for col in new_key_cols))

            if not new_keys:
                return pa.concat_tables([old, new], promote_options="default")

            # Filter old rows whose keys are in new_keys.
            old_missing = [k for k in keys if k not in old.schema.names]
            if old_missing:
                # Old table doesn't have the match_by columns — nothing
                # to match against; just append.
                return pa.concat_tables([old, new], promote_options="default")

            old_key_cols = [old.column(k).to_pylist() for k in keys]
            mask = [
                tuple(col[i] for col in old_key_cols) not in new_keys
                for i in range(old.num_rows)
            ]
            filtered_old = old.filter(pa.array(mask, type=pa.bool_()))
            return pa.concat_tables([filtered_old, new], promote_options="default")

        # OVERWRITE / AUTO / anything else: new wins outright.
        return new


# ---------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------

def _normalize_match_by(match_by: Any) -> tuple[str, ...]:
    """Return *match_by* as a tuple of column names, or ``()`` if unset."""
    if match_by is None or match_by is ...:
        return ()
    if isinstance(match_by, str):
        return (match_by,)
    return tuple(match_by)