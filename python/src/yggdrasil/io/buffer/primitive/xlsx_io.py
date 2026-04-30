"""XLSX I/O for :class:`PrimitiveIO`.

:class:`XlsxIO` handles a single-sheet xlsx workbook. Multi-sheet
workbooks are out of scope at the leaf level — they belong to a
folder/nested IO that maps each sheet onto a child fragment.

Reads use :func:`openpyxl.load_workbook(read_only=True)`; writes
use :func:`openpyxl.Workbook(write_only=True)`. Both modes are
streaming. Save modes: OVERWRITE only.
"""

from __future__ import annotations

import dataclasses
from typing import ClassVar, Iterable, Iterator

import pyarrow as pa
from yggdrasil.data.cast.options import CastOptions
from yggdrasil.data.schema import Schema
from yggdrasil.io.enums import MimeTypes, Mode

from .base import PrimitiveIO

__all__ = ["XlsxIO", "XlsxOptions"]


# ---------------------------------------------------------------------------
# XlsxOptions
# ---------------------------------------------------------------------------


@dataclasses.dataclass(frozen=True, slots=True)
class XlsxOptions(CastOptions):
    """:class:`CastOptions` extended with XLSX-specific knobs."""

    sheet_name: str = "Sheet1"
    has_header: bool = True


# ---------------------------------------------------------------------------
# XlsxIO
# ---------------------------------------------------------------------------


class XlsxIO(PrimitiveIO):
    """:class:`PrimitiveIO` for single-sheet xlsx workbooks."""

    __slots__ = ()

    _FINAL_TABULAR_IO: ClassVar[bool] = True

    @classmethod
    def default_mime_type(cls):
        return MimeTypes.XLSX

    @classmethod
    def options_class(cls):
        return XlsxOptions

    _APPEND_REJECTED_HINT: ClassVar[str] = (
        "XLSX append (whether new rows or new sheets) defeats the "
        "streaming write path — the workbook would need to be reopened "
        "in editable mode and re-saved. Use a folder-oriented writer "
        "to add new xlsx files alongside, or convert to a streamable "
        "format (parquet, jsonl) for incremental ingest."
    )
    _NATIVE_SCANNER_OK: ClassVar[bool] = False

    # ==================================================================
    # Lazy openpyxl import
    # ==================================================================

    @staticmethod
    def _openpyxl():
        try:
            import openpyxl  # noqa: F401
        except ImportError as e:  # pragma: no cover
            raise ImportError(
                "XlsxIO requires openpyxl. Install with: pip install openpyxl"
            ) from e
        import openpyxl
        return openpyxl

    # ==================================================================
    # Schema
    # ==================================================================

    def _collect_schema(self, options: XlsxOptions) -> Schema:
        if self.is_empty():
            return Schema.empty()
        first = next(iter(self._read_arrow_batches(options)), None)
        if first is None:
            return Schema.empty()
        return Schema.from_arrow(first.schema)

    # ==================================================================
    # Read path
    # ==================================================================

    def _read_arrow_batches(
        self,
        options: XlsxOptions,
    ) -> Iterator[pa.RecordBatch]:
        """Stream rows from the named sheet, batch them, yield."""
        if self.is_empty():
            return

        with self._reading_context(options) as io:
            openpyxl = self._openpyxl()
            wb = openpyxl.load_workbook(io, read_only=True, data_only=True)
            try:
                ws = (
                    wb[options.sheet_name]
                    if options.sheet_name in wb.sheetnames
                    else wb[wb.sheetnames[0]]
                )

                row_iter = ws.iter_rows(values_only=True)

                if options.has_header:
                    headers = next(row_iter, None)
                    if headers is None:
                        return
                    columns = [
                        str(h) if h is not None else f"col_{i}"
                        for i, h in enumerate(headers)
                    ]
                else:
                    first_row = next(row_iter, None)
                    if first_row is None:
                        return
                    columns = [f"col_{i}" for i in range(len(first_row))]
                    row_iter = self._chain_one(first_row, row_iter)

                row_size = options.row_size if options.row_size else 4096
                rows: list[dict] = []
                for raw_row in row_iter:
                    rows.append(dict(zip(columns, raw_row)))
                    if len(rows) >= row_size:
                        yield from self._rows_to_batches(rows, options)
                        rows = []
                if rows:
                    yield from self._rows_to_batches(rows, options)
            finally:
                wb.close()

    @staticmethod
    def _chain_one(first, rest):
        yield first
        yield from rest

    def _rows_to_batches(
        self,
        rows: list[dict],
        options: XlsxOptions,
    ) -> Iterator[pa.RecordBatch]:
        normalized = self._normalize_records(rows)
        if not normalized:
            return
        table = pa.Table.from_pylist(normalized)
        for batch in table.to_batches():
            yield options.cast_arrow_tabular(batch)

    # ==================================================================
    # Write path
    # ==================================================================

    def _write_arrow_batches(
        self,
        batches: Iterable[pa.RecordBatch],
        options: XlsxOptions,
    ) -> None:
        """Stream rows into a write-only workbook, save into the buffer.

        openpyxl's ``write_only`` workbook can only ``save(file_like)``
        once — the workbook closes itself on save. We hand it the
        yielded IO directly.
        """
        action = self._resolve_save_mode(options.mode)
        if action is Mode.IGNORE:
            return
        if action is not Mode.OVERWRITE:
            raise NotImplementedError(
                f"{type(self).__name__}._write_arrow_batches only handles "
                f"OVERWRITE; got {action!r}. {self._APPEND_REJECTED_HINT}"
            )

        iterator = iter(batches)
        first = next(iterator, None)
        if first is None:
            return

        if options.target_field is not None:
            first = options.cast_arrow_tabular(first)

        lifecycle = options.copy(truncate_before_write=True)

        with self._writing_context(lifecycle) as io:
            openpyxl = self._openpyxl()
            wb = openpyxl.Workbook(write_only=True)
            try:
                ws = wb.create_sheet(title=options.sheet_name)
                column_names = first.schema.names
                if options.has_header:
                    ws.append(column_names)

                self._append_batch(ws, first, column_names)
                for batch in iterator:
                    if options.target_field is not None:
                        batch = options.cast_arrow_tabular(batch)
                    self._append_batch(ws, batch, column_names)

                io.seek(0)
                wb.save(io)
            finally:
                wb.close()

    @staticmethod
    def _append_batch(ws, batch: pa.RecordBatch, column_names: list[str]) -> None:
        for row in batch.to_pylist():
            ws.append([row.get(c) for c in column_names])