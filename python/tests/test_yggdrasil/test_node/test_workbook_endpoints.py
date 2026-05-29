"""Arrow-IPC tabular preview + workbook (xlsx) endpoints via the live app."""
from __future__ import annotations

import os
import tempfile
import unittest
from pathlib import Path

import pyarrow as pa
import pyarrow.parquet as pq
import pytest

pytest.importorskip("openpyxl")
pytest.importorskip("fastexcel")
import openpyxl  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402

from yggdrasil.node import transport  # noqa: E402
from yggdrasil.node.api.app import create_api  # noqa: E402
from yggdrasil.node.config import Settings  # noqa: E402


def _client(home: Path) -> TestClient:
    return TestClient(create_api(Settings(
        node_id="wb", node_home=home, front_home=home, seed_defaults=False,
    )))


def _seed_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["id", "v"])
    for i in range(5):
        ws.append([i, i * 1.5])
    ws["C2"] = "=SUM(B2:B6)"
    wb.create_sheet("Meta").append(["k", "val"])
    wb.save(str(path))


class TestArrowPreview(unittest.TestCase):
    def test_preview_arrow_decodes(self):
        with tempfile.TemporaryDirectory() as d:
            pq.write_table(pa.table({"id": [1, 2, 3], "city": ["a", "b", "c"]}), os.path.join(d, "t.parquet"))
            c = _client(Path(d))
            r = c.get("/api/v2/tabular/preview.arrow?path=t.parquet&limit=2")
            self.assertEqual(r.status_code, 200)
            self.assertIn("arrow", r.headers["content-type"])
            table = transport.read_arrow_stream(r.content)
            self.assertEqual(table.num_rows, 2)               # bounded
            self.assertEqual(table.column_names, ["id", "city"])


class TestWorkbook(unittest.TestCase):
    def test_sheets_dims(self):
        with tempfile.TemporaryDirectory() as d:
            _seed_xlsx(Path(d) / "b.xlsx")
            c = _client(Path(d))
            js = c.get("/api/v2/workbook/sheets?path=b.xlsx").json()
            names = {s["name"]: s for s in js["sheets"]}
            self.assertEqual(set(names), {"Data", "Meta"})
            self.assertEqual(names["Data"]["rows"], 6)

    def test_read_sheet_arrow_window(self):
        with tempfile.TemporaryDirectory() as d:
            _seed_xlsx(Path(d) / "b.xlsx")
            c = _client(Path(d))
            r = c.get("/api/v2/workbook/read?path=b.xlsx&sheet=Data&n_rows=3")
            self.assertEqual(r.status_code, 200)
            table = transport.read_arrow_stream(r.content)
            self.assertEqual(table.num_rows, 3)

    def test_edit_is_surgical(self):
        with tempfile.TemporaryDirectory() as d:
            p = Path(d) / "b.xlsx"
            _seed_xlsx(p)
            c = _client(Path(d))
            # Cells are stored as sent (numbers stay numbers, =... become
            # formulas) — coercion is the caller's call, like Excel input.
            r = c.post("/api/v2/workbook/edit", json={
                "path": "b.xlsx", "sheet": "Data", "cells": [[2, 2, 99], [3, 2, "=A3*2"]],
            })
            self.assertEqual(r.status_code, 200)
            self.assertEqual(r.json()["cells_written"], 2)
            out = openpyxl.load_workbook(str(p))
            self.assertEqual(out["Data"]["B2"].value, 99)               # number stays number
            self.assertEqual(out["Data"]["B3"].value, "=A3*2")          # new formula stored
            self.assertEqual(out["Data"]["C2"].value, "=SUM(B2:B6)")   # existing formula kept
            self.assertIn("Meta", out.sheetnames)                       # sheet kept

    def test_edit_range(self):
        with tempfile.TemporaryDirectory() as d:
            p = Path(d) / "b.xlsx"
            _seed_xlsx(p)
            c = _client(Path(d))
            r = c.post("/api/v2/workbook/edit", json={
                "path": "b.xlsx", "sheet": "Data",
                "start_row": 2, "start_col": 1, "values": [[100, 200], [300, 400]],
            })
            self.assertEqual(r.json()["cells_written"], 4)
            out = openpyxl.load_workbook(str(p))
            self.assertEqual(out["Data"]["A2"].value, 100)
            self.assertEqual(out["Data"]["B3"].value, 400)

    def test_non_xlsx_rejected(self):
        with tempfile.TemporaryDirectory() as d:
            pq.write_table(pa.table({"x": [1]}), os.path.join(d, "t.parquet"))
            c = _client(Path(d))
            self.assertEqual(c.get("/api/v2/workbook/sheets?path=t.parquet").status_code, 400)


if __name__ == "__main__":
    unittest.main()
